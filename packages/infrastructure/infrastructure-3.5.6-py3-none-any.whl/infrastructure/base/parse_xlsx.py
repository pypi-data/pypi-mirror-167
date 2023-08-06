# -*- coding: utf-8 -*-
# @Author: yongfanmao
# @Date:   2021-03-12 11:38:47
# @E-mail: maoyongfan@163.com
# @Last Modified by:   yongfanmao
# @Last Modified time: 2021-04-29 23:07:43
import datetime
from openpyxl import load_workbook
# linux 导入会报错TypeError: __init__() takes 1 positional argument but 4 were given
from infrastructure.base.con_mysql import UseMysql
from infrastructure.http_agent.http_request import HttpRequest
from infrastructure.hb.hb_request import HBRequest

class ParseXlsx(object):

	def __init__(self,parsePath=""):
		file = load_workbook(parsePath)
		self.parsePath = parsePath
		self.file = file
		self.sheet = file.active 
		#获取sheet页的行数据
		self.rows = self.sheet.rows
		#获取sheet页的列数据
		self.columns = self.sheet.columns

	def readAll(self):
		i = 0
		temp = []
		# 迭代所有的行
		for row in self.rows:
			i = i + 1
			if i == 1:
				continue
			
			line = [col.value for col in row]
			cell_data_1 = self.sheet.cell(row=i, column=1).value #获取第i行1 列的数据
			cell_data_2 = self.sheet.cell(row=i, column=2).value #获取第i行 2 列的数据
			cell_data_4 = self.sheet.cell(row=i, column=4).value #获取第i行 3 列的数据
			# cell_data_4 = self.sheet.cell(row=i, column=18).value #获取第i行 4 列的数据
			temp.append((cell_data_1, cell_data_2, cell_data_4))

		return temp

	def insertServiceInfos(self):
		useMysql = UseMysql()
		nowTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		serviceInfos = self.readAll()
		for serviceInfo in serviceInfos:
			if serviceInfo[2] == "s1":
				level = 1
			else:
				level = 2
			sql = """
				insert into helloBikeDB.helloBikeTools_service12Infos
				(business,service_name,level,
				add_time,change_time)
				Values
				("{business}","{service_name}",
				"{level}",
				"{add_time}","{change_time}")
				""".format(business=serviceInfo[0],
				service_name=serviceInfo[1],
				level=level,
				add_time=nowTime,
				change_time=nowTime)

			print(sql)

			useMysql.executeSql(sql)

	def getServiceTeam(self):
		import time
		tempList =[]
		serviceInfos = self.readAll()
		us = UseMysql()
		headerInfos = us.getTokenInfos()
		helloBikeToken = headerInfos[0]
		user_agent = headerInfos[1]
		hr = HBRequest()
		result =hr.getBussiness(helloBikeToken,user_agent,isServer=False)
		for serviceInfo in serviceInfos:
			if serviceInfo[0] == "服务名":
				continue
			print(serviceInfo[0])
			getTeamUrl = "https://tt.hellobike.cn/v1/api/{service_name}".format(
			service_name=serviceInfo[0])

			data = {"action":"tt.application.info.detail"}

			headers = {'content-type': "application/json;charset=UTF-8",
				'token': helloBikeToken,
				'user-Agent': user_agent}

			response = HttpRequest.post(getTeamUrl,headers=headers,data=data)
			# print(response)

			if response['status']:


				# service_desc = response['result'].get("data").get("desc")
				team = response['result'].get("data").get("team").get("code")
				team_desc = response['result'].get("data").get("team").get("name","")

			
			seachBusiness = False
			for loop in result[1]:
				for teamInfo in result[1][loop]:
					if teamInfo['value'].upper() == team.upper():
						business = loop
						seachBusiness = True
						break
				if seachBusiness:
					break
			print(team,business)
			tempList.append([team,business])
			# time.sleep(3)

		a = ParseXlsx(parsePath="/Users/yongfanmao/Desktop/a.xlsx")
		a.insertXlsx(tempList)


	def insertXlsx(self,resultList):
		
		for data in resultList:
			self.sheet.append(data)
			self.file.save(self.parsePath)


if __name__ == '__main__':
	# a = ParseXlsx("/Users/yongfanmao/Downloads/服务列表.xlsx")
	# print(a.insertServiceInfos())
	a = ParseXlsx(parsePath="/Users/yongfanmao/Downloads/jacoco_app.xlsx")
	a.getServiceTeam()


